#!/usr/bin/env python3
"""
append_to_atlas.py — persist scored pain points into the Boring Pain Atlas.

Writes to BOTH boring_pain_atlas.xlsx (human-sortable) and boring_pain_atlas.jsonl
(machine / BigQuery-ready). Idempotent: rows dedup on a hash of (sub_vertical|pain_point),
so re-scanning a vertical UPDATES rows instead of duplicating them.

Usage
-----
Append/update rows from a JSON file (a list of row dicts; see references/atlas-schema.md):
    python append_to_atlas.py --rows rows.json [--atlas /path/to/boring_pain_atlas.xlsx]
                              [--template /path/to/atlas_template.xlsx]

Query the existing atlas without re-researching:
    python append_to_atlas.py --atlas <path> --query --min-composite 24
    python append_to_atlas.py --atlas <path> --query --sector "Healthcare" --category "Compliance/Risk"
    python append_to_atlas.py --atlas <path> --query --top 15

Notes
-----
- drudge_composite is ALWAYS recomputed as the sum of the six sub-scores (any supplied value is ignored).
- A row with an empty evidence_source_1 is rejected — the library stays honest.
- pain_category must be one of the five enums.
"""

import argparse
import datetime as _dt
import hashlib
import json
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl --break-system-packages")

# --- Field contract (must match references/atlas-schema.md, in order) ---
FIELDS = [
    "atlas_id", "date_scanned", "sector", "sub_vertical", "segment",
    "pain_point", "description", "pain_category", "boring_task", "root_cause",
    "target_persona", "current_workaround", "frequency_note", "annual_impact_usd",
    "drudge_D_digital", "drudge_R_repetitive", "drudge_U_uniform",
    "drudge_D_data", "drudge_G_grievous", "drudge_E_expensive", "drudge_composite",
    "drudge_cap_flag", "viability_wedge", "viability_tam", "viability_incumbency",
    "viability_regmoat", "viability_note", "agent_concept",
    "evidence_source_1", "evidence_tier", "onet_soc_codes", "confidence", "notes",
]

DRUDGE_SUBSCORES = [
    "drudge_D_digital", "drudge_R_repetitive", "drudge_U_uniform",
    "drudge_D_data", "drudge_G_grievous", "drudge_E_expensive",
]
VIABILITY_SUBSCORES = [
    "viability_wedge", "viability_tam", "viability_incumbency", "viability_regmoat",
]
CATEGORIES = {"Financial", "Productivity", "Process", "Support", "Compliance/Risk"}
CONFIDENCE = {"High", "Med", "Low"}


def make_id(sub_vertical: str, pain_point: str) -> str:
    key = f"{(sub_vertical or '').strip().lower()}|{(pain_point or '').strip().lower()}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]


def validate_and_normalize(row: dict) -> dict:
    """Validate one row, fill defaults, compute composite. Raises ValueError on bad input."""
    out = {f: row.get(f, "") for f in FIELDS}

    if not out["sub_vertical"]:
        raise ValueError("sub_vertical is required")
    if not out["pain_point"]:
        raise ValueError("pain_point is required")
    if not str(out["evidence_source_1"]).strip():
        raise ValueError(f"evidence_source_1 is required (row '{out['pain_point']}') — no source, no row")

    if out["pain_category"] not in CATEGORIES:
        raise ValueError(
            f"pain_category '{out['pain_category']}' invalid for '{out['pain_point']}'. "
            f"Must be one of: {sorted(CATEGORIES)}"
        )

    if out["confidence"] and out["confidence"] not in CONFIDENCE:
        raise ValueError(f"confidence '{out['confidence']}' must be one of {sorted(CONFIDENCE)}")

    # DRUDGE sub-scores must be ints 1-5; composite recomputed.
    total = 0
    for f in DRUDGE_SUBSCORES:
        try:
            v = int(out[f])
        except (TypeError, ValueError):
            raise ValueError(f"{f} must be an integer 1-5 (row '{out['pain_point']}')")
        if not 1 <= v <= 5:
            raise ValueError(f"{f}={v} out of range 1-5 (row '{out['pain_point']}')")
        out[f] = v
        total += v
    out["drudge_composite"] = total

    # Viability sub-scores: ints 1-5 if provided (allowed blank for low-DRUDGE rows).
    for f in VIABILITY_SUBSCORES:
        if out[f] not in ("", None):
            try:
                v = int(out[f])
            except (TypeError, ValueError):
                raise ValueError(f"{f} must be an integer 1-5 or blank (row '{out['pain_point']}')")
            if not 1 <= v <= 5:
                raise ValueError(f"{f}={v} out of range 1-5 (row '{out['pain_point']}')")
            out[f] = v

    if out["evidence_tier"] not in ("", None):
        try:
            out["evidence_tier"] = int(out["evidence_tier"])
        except (TypeError, ValueError):
            raise ValueError(f"evidence_tier must be int 1-4 or blank (row '{out['pain_point']}')")

    out["atlas_id"] = make_id(out["sub_vertical"], out["pain_point"])
    out["date_scanned"] = out["date_scanned"] or _dt.date.today().isoformat()
    return out


def load_or_create(atlas_path: str, template_path: str):
    if os.path.exists(atlas_path):
        wb = load_workbook(atlas_path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        if header != FIELDS:
            raise ValueError("Existing atlas header does not match the schema; refusing to write.")
        return wb, ws
    if template_path and os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb.active
        return wb, ws
    # No file, no template: create a minimal workbook with the header row.
    wb = Workbook()
    ws = wb.active
    ws.title = "atlas"
    ws.append(FIELDS)
    ws.freeze_panes = "A2"
    return wb, ws


def existing_index(ws):
    """Map atlas_id -> row number for dedup/update."""
    idx = {}
    for r in range(2, ws.max_row + 1):
        aid = ws.cell(row=r, column=1).value
        if aid:
            idx[aid] = r
    return idx


def append_rows(rows, atlas_path, template_path):
    parent = os.path.dirname(os.path.abspath(atlas_path))
    os.makedirs(parent, exist_ok=True)
    wb, ws = load_or_create(atlas_path, template_path)
    idx = existing_index(ws)

    normalized = [validate_and_normalize(r) for r in rows]  # validate all before writing

    added, updated = 0, 0
    for row in normalized:
        values = [row[f] for f in FIELDS]
        if row["atlas_id"] in idx:
            r = idx[row["atlas_id"]]
            for col, val in enumerate(values, start=1):
                ws.cell(row=r, column=col, value=val)
            updated += 1
        else:
            ws.append(values)
            idx[row["atlas_id"]] = ws.max_row
            added += 1

    wb.save(atlas_path)

    # JSONL twin: rewrite from the workbook so it always mirrors the xlsx exactly.
    jsonl_path = os.path.splitext(atlas_path)[0] + ".jsonl"
    with open(jsonl_path, "w", encoding="utf-8") as fh:
        for r in range(2, ws.max_row + 1):
            obj = {FIELDS[c - 1]: ws.cell(row=r, column=c).value for c in range(1, len(FIELDS) + 1)}
            fh.write(json.dumps(obj, ensure_ascii=False) + "\n")

    return added, updated, atlas_path, jsonl_path


def query(atlas_path, min_composite=None, sector=None, category=None, top=None):
    if not os.path.exists(atlas_path):
        sys.exit(f"No atlas at {atlas_path}")
    wb = load_workbook(atlas_path, read_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        obj = {FIELDS[c - 1]: ws.cell(row=r, column=c).value for c in range(1, len(FIELDS) + 1)}
        if obj.get("atlas_id") is None:
            continue
        if min_composite is not None and (obj.get("drudge_composite") or 0) < min_composite:
            continue
        if sector and (obj.get("sector") or "").lower() != sector.lower():
            continue
        if category and (obj.get("pain_category") or "") != category:
            continue
        rows.append(obj)
    rows.sort(key=lambda o: o.get("drudge_composite") or 0, reverse=True)
    if top:
        rows = rows[:top]
    for o in rows:
        print(f"[{o.get('drudge_composite'):>2}] {o.get('sub_vertical')} :: {o.get('pain_point')} "
              f"({o.get('pain_category')}) — {o.get('agent_concept')}")
    print(f"\n{len(rows)} row(s).")


def main():
    ap = argparse.ArgumentParser(description="Append/query the Boring Pain Atlas.")
    ap.add_argument("--atlas", default="/mnt/user-data/outputs/boring_pain_atlas.xlsx")
    ap.add_argument("--template", default=os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "assets", "atlas_template.xlsx"))
    ap.add_argument("--rows", help="Path to a JSON file containing a list of row dicts.")
    ap.add_argument("--query", action="store_true")
    ap.add_argument("--min-composite", type=int, default=None)
    ap.add_argument("--sector", default=None)
    ap.add_argument("--category", default=None)
    ap.add_argument("--top", type=int, default=None)
    args = ap.parse_args()

    if args.query:
        query(args.atlas, args.min_composite, args.sector, args.category, args.top)
        return

    if not args.rows:
        ap.error("provide --rows <file.json> to append, or --query to read.")

    with open(args.rows, encoding="utf-8") as fh:
        data = json.load(fh)
    if isinstance(data, dict):
        data = [data]

    added, updated, xlsx_p, jsonl_p = append_rows(data, args.atlas, args.template)
    print(f"Appended {added}, updated {updated}.")
    print(f"  xlsx : {xlsx_p}")
    print(f"  jsonl: {jsonl_p}")


if __name__ == "__main__":
    main()
